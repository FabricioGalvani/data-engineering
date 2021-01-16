# Reading an excel file using Python
from openpyxl import load_workbook
import xlsxwriter
import csv


def load_csv_files(input_file):
    '''
    Function to load the csv files.

    Parameters:
        input_file (str): File input.
    
    Returns:
        file (obj): The csv file.
    '''
    with open(input_file, encoding='utf-8') as csvfile:
        file = csv.reader(csvfile, delimiter=';')

        if file:
            return file 
        else:
            return False


def load_xlsx_files(input_file):
    '''
    Function to load the xlsx files..

    Parameters:
        input_file (str): File input.
    
    Returns:
        sheet (obj): The sheet.
    '''

    wb = load_workbook(input_file)

    sheet = wb['Processed']

    if sheet:
        return sheet
    else:
        return False


def format_date(transaction_date):
    '''
    Function to format the date.

    Parameters:
        transaction_date (date): The date which we will format.
    
    Returns:
        transaction_date (str): The formatted date(YYYYMMDD).
    '''

    if transaction_date:
        transaction_date = transaction_date.date()
        transaction_date = transaction_date.strftime('%Y%m%d')

    return transaction_date


def format_cost(total_cost):
    '''
    Function to add the percentage to the value.

    Parameters:
        total_cost (int): The value which we will add the value.
    
    Returns:
        total_cost (float): The added value of the increase.
    '''

    percentage = 1.10
    total_cost = round(float(total_cost) * percentage, 2)
    
    return total_cost


def format_cost_center(cost_center_name):
    '''
    Function to format the name of the cost center.

    Parameters:
        cost_center_name (str): The cost center name.
    
    Returns:
        cost_center_name (str): The name of the cost center formatted if necessary.
    '''

    if cost_center_name == "Faturamento":
        cost_center_name = "Contas a Pagar/Receber"

    return cost_center_name


def transform_data(sheet):
    '''
    Function to format data.

    Parameters:
        sheet (obj): The file object.
    
    Returns:
        total_cost (list): The list of data.
    '''

    total_cost = []
    
    rows = sheet.iter_rows(min_row=1, max_row=1)
    first_row = next(rows)
    header = [data.value for data in first_row]
    total_cost.append(header)

    for row in sheet.iter_rows(values_only = True, min_row=2):
        row = list(row)
        row[1] = format_date(row[1])
        row[2] = format_cost_center(row[2])
        row[3] = format_cost(row[3])
        
        total_cost.append(row)
    
    return total_cost


def save_sheet(file):
    '''
    Function to save the file.

    Parameters:
        file (str): The file to save.
    
    Returns:
        True (bool): Return Boolean expression.
    '''
    
    with xlsxwriter.Workbook('TotalCost.xlsx') as workbook:
        worksheet = workbook.add_worksheet()

        for row_num, data in enumerate(total_cost):
            worksheet.write_row(row_num, 0, data)

    return True


if __name__ == "__main__":
    csv_input_file = "files/CostCenter.csv"
    load_csv_files(csv_input_file)
    
    xlsx_input_file = "files/Values.xlsx"
    sheet = load_xlsx_files(xlsx_input_file)
    
    if sheet:
        total_cost = transform_data(sheet)
        save_sheet(total_cost)
